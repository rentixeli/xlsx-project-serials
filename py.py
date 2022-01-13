#!/usr/bin/python3
import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, Color, Fill, PatternFill, GradientFill, Border, Side, Alignment, Protection
from openpyxl.workbook import Workbook

print('''
██╗  ██╗██╗     ███████╗██████╗ ██████╗ 
╚██╗██╔╝██║     ██╔════╝╚════██╗██╔══██╗
 ╚███╔╝ ██║     ███████╗ █████╔╝██████╔╝
 ██╔██╗ ██║     ╚════██║ ╚═══██╗██╔══██╗
██╔╝ ██╗███████╗███████║██████╔╝██║  ██║
╚═╝  ╚═╝╚══════╝╚══════╝╚═════╝ ╚═╝  ╚═╝
	by: Rentix Eliyahu
''')

def main():
	print ("Welcome to xls3r-maker. What would you like to do?")
	print ("1) PRE (C50).")
	print ("2) POST (B50).")
	print ("3) Both.")
	print ('4) Exit.')
	try:
		mChoice = int(input("My answer is: "))
	except:
		print ("Not a number...")
		main()

	if mChoice == 1:	# choice number one - all files
		filelist = []
		for file in glob.glob("*.xlsx"):
			filelist.append(file)
		#print (filelist)

		values = [] # list of all the B8 values he gets from all files.
		wos = [] # list of all working sheets
		for file in filelist:
			wf = openpyxl.load_workbook(file)
			sheet_number = len(wf.worksheets)
			for ws in wf.worksheets:
				wos.append(ws.title) # wos = list of all the sheets.

			ws = wf[wos[0]]	# wos[2] 3rd sheet.
			sheet = ws['C10']._value
			values.append(sheet)
		#print (values)

		pre = Workbook()
		dest = 'pre.xlsx'
		ws1 = pre.active
		ws1.title="Sheet1"
		ws1['A1'] = "Serial Number"
		ws1['B1'] = "Main GL"
		ws1['C1'] = "*16"
		ws1['D1'] = "Blue"
		num = 0
		for i in range(2, len(filelist)+2):
			cell = 'A' + str(i)
			filenamm = filelist[num]
			ws1[cell] = filenamm[:-5]
			num += 1
		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'B' + str(i)
			ws1[cell] = values[num]
			summ += int(values[num])
			num += 1
		cell = 'B' + str(len(values)+2)
		ws1[cell] = summ / len(values)

		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'C' + str(i)
			dub = int(values[num]) * 16
			ws1[cell] = dub
			summ += dub
			num += 1

		cell = 'C' + str(len(values)+2)
		ws1[cell] = summ / len(values)


		dub = 0
		for i in range(2, len(values)+2):
			cell = 'D' + str(i)
			nam = 'C' + str(i)
			val = ws1[nam]._value
			nam2 = 'C' + str(len(values)+2)
			val2 = ws1[nam2]._value
			dub = val - val2
			dub = dub / val2
			dub = dub * 100

			ws1[cell] = dub
			num += 1
		pre.save(filename = dest)

	elif mChoice == 2:
		filelist = []
		for file in glob.glob("*.xlsx"):
			filelist.append(file)
		#print (filelist)

		values = [] # list of all the B8 values he gets from all files.
		wos = [] # list of all working sheets
		for file in filelist:
			wf = openpyxl.load_workbook(file)
			sheet_number = len(wf.worksheets)
			for ws in wf.worksheets:
				wos.append(ws.title) # wos = list of all the sheets.

			ws = wf[wos[1]]	# wos[2] 3rd sheet.
			sheet = ws['C10']._value
			values.append(sheet)
		#print (values)

		post = Workbook()
		dest = 'post.xlsx'
		ws1 = post.active
		ws1.title="Sheet1"
		ws1['A1'] = "Serial Number"
		ws1['B1'] = "Main GL"
		ws1['C1'] = "*16"
		ws1['D1'] = "Blue"
		num = 0
		for i in range(2, len(filelist)+2):
			cell = 'A' + str(i)
			filenamm = filelist[num]
			ws1[cell] = filenamm[:-5]
			num += 1
		num = 0
		for i in range(2, len(values)+2):
			cell = 'B' + str(i)
			ws1[cell] = values[num]
			num += 1

		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'C' + str(i)
			dub = int(values[num]) * 16
			ws1[cell] = dub
			summ += dub
			num += 1

		cell = 'C' + str(len(values)+2)
		ws1[cell] = summ / len(values)


		dub = 0
		for i in range(2, len(values)+2):
			cell = 'D' + str(i)
			nam = 'C' + str(i)
			val = ws1[nam]._value
			nam2 = 'C' + str(len(values)+2)
			val2 = ws1[nam2]._value
			dub = val - val2
			dub = dub / val2
			dub = dub * 100

			ws1[cell] = dub
			num += 1
		post.save(filename = dest)

	elif mChoice == 3:
		filelist = []
		for file in glob.glob("*.xlsx"):
			filelist.append(file)
		#print (filelist)

		values = [] # list of all the B8 values he gets from all files.
		values2 = []
		wos = [] # list of all working sheets
		for file in filelist:
			wf = openpyxl.load_workbook(file)
			sheet_number = len(wf.worksheets)
			for ws in wf.worksheets:
				wos.append(ws.title) # wos = list of all the sheets.

			ws = wf[wos[0]]	# wos[2] 3rd sheet.
			pre = ws['C10']._value
			values.append(pre)

			ws = wf[wos[1]]
			post = ws['C10']._value
			values2.append(post)
		#print (values)

		alls = Workbook()
		dest = 'ALL.xlsx'
		ws1 = alls.active
		ws1.title="Sheet1"
		ws1['A1'] = "Serial Number"
		ws1['B1'] = "Pre"
		ws1['C1'] = "Pre Green"
		ws1['D1'] = "Pre Blue"
		ws1['E1'] = "Post"
		ws1['F1'] = "Post Green"
		ws1['G1'] = "Post Blue"
		num = 0
		for i in range(2, len(filelist)+2):
			cell = 'A' + str(i)
			filenamm = filelist[num]
			ws1[cell] = filenamm[:-5]
			num += 1
		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'B' + str(i)
			ws1[cell] = values[num]
			summ += int(values[num])
			num += 1
		cell = 'B' + str(len(values)+2)
		ws1[cell] = summ / len(values)

		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'C' + str(i)
			dub = int(values[num]) * 16
			ws1[cell] = dub
			summ += dub
			num += 1

		cell = 'C' + str(len(values)+2)
		ws1[cell] = summ / len(values)


		dub = 0
		for i in range(2, len(values)+2):
			cell = 'D' + str(i)
			nam = 'C' + str(i)
			val = ws1[nam]._value
			nam2 = 'C' + str(len(values)+2)
			val2 = ws1[nam2]._value
			dub = val - val2
			dub = dub / val2
			dub = dub * 100

			ws1[cell] = dub
			num += 1

		num = 0
		for i in range(2, len(values)+2):
			cell = 'E' + str(i)
			ws1[cell] = values[num]
			num += 1

		num = 0
		summ = 0
		for i in range(2, len(values)+2):
			cell = 'F' + str(i)
			dub = int(values[num]) * 16
			ws1[cell] = dub
			summ += dub
			num += 1

		cell = 'F' + str(len(values)+2)
		ws1[cell] = summ / len(values)


		dub = 0
		for i in range(2, len(values)+2):
			cell = 'G' + str(i)
			nam = 'F' + str(i)
			val = ws1[nam]._value
			nam2 = 'F' + str(len(values)+2)
			val2 = ws1[nam2]._value
			dub = val - val2
			dub = dub / val2
			dub = dub * 100

			ws1[cell] = dub
			num += 1
		alls.save(filename = dest)

	elif mChoice == 4:
		sys.exit()

	else:
		print('I don\'t know.. try again..')
		main()






main()
